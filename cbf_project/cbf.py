import pandas as pd
import string
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def set_fonts(ws):
    title_font = Font(size=16, bold=True)
    table_head_font = Font(size=14)
    write_font = Font(size=12)
    perc_font = Font(size=12, bold=True)
    
    ws['A4'].font, ws['A5'].font, ws['A6'].font = [title_font] * 3
    ws['E11'].font, ws['E12'].font, ws['E13'].font = [table_head_font] * 3
    ws['E11'].font, ws['E12'].font, ws['E13'].font = [table_head_font] * 3
    ws['G11'].font, ws['G12'].font, ws['G13'].font = [table_head_font] * 3
    ws['I13'].font, ws['K13'].font = [table_head_font] * 2
    
    ws['A23'].font, ws['A24'].font = [write_font] * 2
    ws['A15'].font, ws['E15'].font, ws['G15'].font, ws['I15'].font = [write_font] * 4
    ws['A16'].font, ws['E16'].font, ws['G16'].font, ws['I16'].font = [write_font] * 4
    ws['A17'].font, ws['E17'].font, ws['G17'].font, ws['I17'].font = [write_font] * 4
    
    ws['K15'].font, ws['K17'].font, ws['K16'].font = [perc_font]*3
    ws['A20'].font = Font(size = 10)
    
def set_align(ws):
    num_align = Alignment(horizontal='right')
    
    ws['E11'].alignment, ws['E12'].alignment, ws['E13'].alignment, ws['E15'].alignment, ws['E17'].alignment = [num_align]*5     
    ws['G11'].alignment, ws['G12'].alignment, ws['G13'].alignment, ws['G15'].alignment, ws['g17'].alignment = [num_align]*5
    ws['I13'].alignment, ws['I15'].alignment, ws['I17'].alignment = [num_align]*3
    ws['K13'].alignment, ws['K15'].alignment, ws['K17'].alignment = [num_align]*3
    
    ws['E16'].alignment, ws['G16'].alignment, ws['I16'].alignment, ws['K16'].alignment = [num_align]*4
    


df = pd.read_csv("cbf2.csv")
drop_cols = ['ID','Pastor','City','State','Zip','Unnamed: 15','Unnamed: 16']

df['mrp_perc_diff'] = df['mrp_perc_diff'].fillna(0)
df['cbf_perc_diff'] = df['cbf_perc_diff'].fillna(0)
df = df.dropna()

a5 = "Comparison of Giving"
a6 = "Fiscal Years Ending 3/31/17 and 3/31/16"
e11_g11 = "Fiscal Year"
e12_g12 = "Ended"
e13, g13 = "3/31/17", "3/31/16"
i13, k13 = "$ Change", "% Change"
a20_mrp = "*CBFNC Contributions may include MRP allocation, Mission & Ministry Offering and other direct gifts."
a23 = "Overall, contributions to CBFNC from churches decreased 5.98% for this period."
a24 = "If you have questions, please call Gail McAlister, Financial Manager, CBFNC at 336-759-3456."
a15_mrp, a15 = "CBFNC Contributions *", "CBFNC Contributions"

wb = Workbook()
i = 0
for row in df.iterrows():
#     print cname,
    r = row[1]
    mrp = 0 if r['MRP-Flag'] == 0 else 1
    cname = r['Name']
    cbf_new, cbf_old = r['cbf_new'], r['cbf_old']
    cbf_diff, cbf_perc_diff = r['cbf_diff'], r['cbf_perc_diff']
    ws = wb.create_sheet(cname[:30], i)
    
    set_fonts(ws)
    set_align(ws)
    
    # Build basics of worksheet
    ws['A4'] = cname
    ws['A5'] = a5
    ws['A6'] = a6
    ws['E11'], ws['G11'] = e11_g11, e11_g11
    ws['E12'], ws['G12'] = e12_g12, e12_g12
    ws['E13'], ws['G13'] = e13, g13
    ws['I13'], ws['K13'] = i13, k13
    ws['A23'], ws['A24'] = a23, a24
    
    if mrp:
        # finish building an MRP worksheet
        ws['A15'] = "MRP Contributions"
        ws['A17'] = a15_mrp
        ws['A20'] = a20_mrp
        
        # set numeric mrp vals
        ws['E15'], ws['G15'] = "$"+str(r['mrp_new']), "$"+str(r['mrp_old'])
        ws['I15'], ws['K15'] = "$"+str(r['mrp_diff']), str(r['mrp_perc_diff'])+"%"
        
        # set numeric cbf_vals
        ws['E17'], ws['G17'] = "$"+str(r['cbf_new']), "$"+str(r['cbf_old'])
        ws['I17'], ws['K17'] = "$"+str(r['cbf_diff']), str(r['cbf_perc_diff'])+"%"

    else:
        # Build a non-MRP worksheet
        # set numeric cbf_vals
        ws['E16'], ws['G16'] = "$"+str(r['cbf_new']), "$"+str(r['cbf_old'])
        ws['I16'], ws['K16'] = "$"+str(r['cbf_diff']), str(r['cbf_perc_diff'])+"%"
        
        ws['A16'] = a15
    i += 1
    
wb.save(filename="cbf_report.xlsx")



